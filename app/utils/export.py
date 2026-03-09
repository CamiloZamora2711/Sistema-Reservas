from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from io import BytesIO
from datetime import datetime

def exportar_excel(reservas, fecha_inicio, fecha_fin):
    """Exportar reservas a Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reservas"
    
    # Título
    ws.merge_cells('A1:G1')
    ws['A1'] = f'Reporte de Reservas - {fecha_inicio} a {fecha_fin}'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Encabezados
    headers = ['ID', 'Sala', 'Usuario', 'Fecha', 'Bloque', 'Estado', 'Motivo']
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    for row, reserva in enumerate(reservas, start=4):
        ws.cell(row=row, column=1, value=reserva.id)
        ws.cell(row=row, column=2, value=reserva.sala.nombre)
        ws.cell(row=row, column=3, value=reserva.usuario.nombre)
        ws.cell(row=row, column=4, value=reserva.fecha.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=5, value=reserva.bloque)
        ws.cell(row=row, column=6, value=reserva.estado.upper())
        ws.cell(row=row, column=7, value=reserva.motivo)
    
    # Ajustar anchos de columna
    column_widths = [8, 30, 25, 12, 15, 12, 40]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

def exportar_pdf(reservas, fecha_inicio, fecha_fin):
    """Exportar reservas a PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Título
    title = Paragraph(f'Reporte de Reservas<br/>{fecha_inicio} a {fecha_fin}', title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Información general
    info_style = styles['Normal']
    info = Paragraph(f'<b>Total de reservas:</b> {len(reservas)}<br/>'
                    f'<b>Fecha de generación:</b> {datetime.now().strftime("%d/%m/%Y %H:%M")}',
                    info_style)
    elements.append(info)
    elements.append(Spacer(1, 0.3*inch))
    
    # Tabla de datos
    data = [['Sala', 'Usuario', 'Fecha', 'Bloque', 'Estado']]
    
    for reserva in reservas:
        data.append([
            reserva.sala.nombre,
            reserva.usuario.nombre,
            reserva.fecha.strftime('%d/%m/%Y'),
            reserva.bloque,
            reserva.estado.upper()
        ])
    
    # Crear tabla
    table = Table(data, colWidths=[2*inch, 2*inch, 1*inch, 1.2*inch, 1*inch])
    
    # Estilo de tabla
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    elements.append(table)
    
    # Generar PDF
    doc.build(elements)
    buffer.seek(0)
    
    return buffer
